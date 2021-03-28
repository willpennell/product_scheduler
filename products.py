import openpyxl, os
class NewProduct:
    """ class for product input """
    def __init__(self):
        """ Initailise new product object """
        self.labels = { 'apollo sound': 'Apollo Sound', 'delectable records': 'Delectable Records', 'element one': 'Element One', 'get down samples': 'Get Down Samples',
            'house of loop': 'House of Loop', 'lp24': 'LP24', 'mootant': 'Mootant', 'mind flux': 'MIND FLUX', 'odd smpls': 'ODD Smpls', 'shuriken audio': 'Shuriken Audio',
            'soul rush': 'Soul Rush', 'system 6 samples': 'System 6 Samples', 'samplesound': 'Samplesound', 'samplestate': 'SAMPLESTATE', 'unity records': 'UNITY RECORDS',
            '91 vocals': '91 Vocals', 'aubit sounds': 'Aubit Sounds', 'audeo box': 'Audeo Box', 'audiomodern': 'Audiomodern', 'audiostrasse': 'Audiostrasse', 'big fish audio': 'Big Fish Audio',
            'black octopus': 'Black Octopus', 'bingoshakerz': 'Bingoshakerz', 'dropgun audio': 'DropGun Audio', 'datacode': 'Datacode', 'engineering samples': 'Engineering Samples',
            'image sounds': 'Image Sounds', 'komorebi audio': 'Komorebi Audio', 'laniakea sounds': 'Laniakea Sounds', 'mode audio': 'ModeAudio', 'ost audio': 'OST Audio', 'prime loops': 'Prime Loops',
            'pelham and junior': 'Pelham & Junior', 'production master': 'Production Master', 'producer loops': 'Producer Loops', 'resonance sound': 'Resonance Sound', 'riemann': 'Riemann',
            'samplestar': 'Samplestar', 'soundsmiths': 'Soundsmiths', 'shamanstems': 'ShamanStems', 'skifonix': 'Skifonix', 'toolroom': 'Toolroom', 'triad': 'Triad', 'zenhiser': 'Zenhiser' 
            }
        os.chdir('C:\\Users\\willp\\PythonScripts\\Schedule')
    
    def data_input(self):
        """ Input for new product """
        self.date = str(input('Enter Date:'))
        self.label = input('Enter Label: ')
        self.product = input('Enter Product: ')

    def open_workbook(self):
        """ Open product schedule """
        self.releaseWB = openpyxl.load_workbook('WillsReleases.xlsx')

    def are_you_finished(self):
        """finishes loop """
        self.flag = input('Are you finished? ')
        if self.flag == 'y':
            return self.flag
        
    def add_product_to_workbook_date_not_known(self):
        """ adds product to nexr avaliable date """
        labelSheet = self.releaseWB[self.labels[self.label]]
        self.start = 0
        self.date = input('Enter this weeks release date to add new product to next aviliable slot: ')
        for i, cellObj in enumerate(list(labelSheet.columns)[0]):
            if cellObj.value == self.date:
                self.start = i + 1
                break
        for j in range(self.start, labelSheet.max_row):
            if labelSheet.cell(row=j+1, column=2).value == None:
                labelSheet.cell(row=j+1, column=2).value = self.product
                break

    def add_product_to_workbook_date_known(self):
        """ add new product to workbook """
        labelSheet = self.releaseWB[self.labels[self.label]]
        for i, cellObj in enumerate(list(labelSheet.columns)[0]):
            if cellObj.value == self.date:
                labelSheet.cell(row=i+1, column=2).value = self.product
                self.releaseWB.save('WillsReleases.xlsx')
            
    def weekly_release_generator(self):
        """ generates weekly release list for product set up """
        weeklydoc = open(f'{self.date}_Wills_Release.txt', 'a')
        weeklydoc.write(f'{self.date} Releases\n\n')
        for labelpage in self.releaseWB.sheetnames:
            labelsheet = self.releaseWB[labelpage]
            for i, cellObj in enumerate(list(labelsheet.columns)[0]):
                if cellObj.value == self.date:
                    if labelsheet.cell(row=i+1, column=2).value != None:
                        release = labelsheet.cell(row=i+1, column=2).value
                        weeklydoc.write(f"{release}[{labelsheet.title}]\n")
        weeklydoc.close()


    def close_workbook(self):
        """ Save and close workbook """
        self.releaseWB.save('WillsReleases.xlsx')
        self.releaseWB.close()

    
    

        